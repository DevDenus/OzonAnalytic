import io
from typing import Tuple, List

from openpyxl import Workbook

from index_db.operations import BrandRepository, SellerRepository, ProductRepository

def get_sellers_names(db) -> List[str]:
    sellers = SellerRepository.get_all(db)
    sellers_names = [seller.name for seller in sellers]
    return sellers_names

def get_brands_names(db) -> List[str]:
    brands = BrandRepository.get_all(db)
    brands_names = [brand.name for brand in brands]
    return brands_names

def get_product_count(db) -> int:
    return ProductRepository.get_product_count(db)

def make_seller_report(seller_name : str, db) -> Tuple[io.BytesIO, str]:
    seller = SellerRepository.get_by_name(db, seller_name)
    if seller is None:
        raise KeyError(f"Seller {seller_name} does not exist!")
    seller_id = seller.id

    workbook = Workbook()
    workbook_state = workbook.active
    workbook_state.title = f"Отчёт по {seller_name}"

    workbook_state.append([
        "Имя товара", "Артикул", "Ссылка", "Брэнд", "Цена по Ozon карте",
        "Цена", "Рейтинг", "Количество отзывов", "Количество вопросов", "На распродаже"
    ])

    products_states = ProductRepository.get_by_seller_id(db, seller_id)
    for product, product_state in products_states:
        product_name = product.name
        product_pk = product.pk
        product_url = product.url
        product_brand = BrandRepository.get_by_id(db, product.brand_id)
        if product_brand is None:
            product_brand_name = ""
        else:
            product_brand_name = product_brand.name
        product_price_ozon = product_state.price_ozon_card
        product_price = product_state.price
        product_rating = product_state.rating
        product_reviews = product_state.review_count
        product_questions = product_state.question_count
        product_on_sale = product_state.on_sale
        workbook_state.append([
            product_name, product_pk, product_url, product_brand_name,
            product_price_ozon, product_price, product_rating,
            product_reviews, product_questions, product_on_sale
        ])

    report_file = io.BytesIO()
    workbook.save(report_file)
    report_file.seek(0)
    file_name = f"{seller_name}.xlsx"
    return report_file, file_name

def make_brand_report(brand_name : str, db) -> Tuple[io.BytesIO, str]:
    brand = BrandRepository.get_by_name(db, brand_name)
    if brand is None:
        raise KeyError(f"Brand {brand_name} does not exist!")
    brand_id = brand.id

    workbook = Workbook()
    workbook_state = workbook.active
    workbook_state.title = f"Отчёт по {brand_name}"

    workbook_state.append([
        "Имя товара", "Артикул", "Ссылка", "Продавец", "Цена по Ozon карте",
        "Цена", "Рейтинг", "Количество отзывов", "Количество вопросов", "На распродаже"
    ])

    products_states = ProductRepository.get_by_brand_id(db, brand_id)
    for product, product_state in products_states:
        product_name = product.name
        product_pk = product.pk
        product_url = product.url
        product_seller_id = product.seller_id
        if product_seller_id:
            product_seller = SellerRepository.get_by_id(db, product_seller_id).name
        else:
            product_seller = ""
        product_price_ozon = product_state.price_ozon_card
        product_price = product_state.price
        product_rating = product_state.rating
        product_reviews = product_state.review_count
        product_questions = product_state.question_count
        product_on_sale = product_state.on_sale
        workbook_state.append([
            product_name, product_pk, product_url, product_seller,
            product_price_ozon, product_price, product_rating,
            product_reviews, product_questions, product_on_sale
        ])

    report_file = io.BytesIO()
    workbook.save(report_file)
    report_file.seek(0)
    file_name = f"{brand_name}.xlsx"
    return report_file, file_name

def make_product_report(product_pk : int, db) -> Tuple[io.BytesIO, str]:
    product = ProductRepository.get_by_pk(db, product_pk)
    if product is None:
        raise KeyError(f"Product with {product_pk} does not exist!")

    workbook = Workbook()
    workbook_state = workbook.active
    workbook_state.title = f"Отчёт по товару {product_pk}"

    workbook_state.append([
        "Цена по Ozon карте", "Цена", "Рейтинг", "Количество отзывов",
        "Количество вопросов", "На распродаже", "Дата"
    ])

    product_history = ProductRepository.get_product_history(db, product.id)
    for state in product_history:
        price_ozon = state.price_ozon_card
        price = state.price
        rating = state.rating
        reviews = state.review_count
        questions = state.question_count
        on_sale = state.on_sale
        created_at = state.created_at
        workbook_state.append([
            price_ozon, price, rating, reviews, questions, on_sale, created_at
        ])

    report_file = io.BytesIO()
    workbook.save(report_file)
    report_file.seek(0)
    file_name = f"{product_pk}.xlsx"
    return report_file, file_name

def make_products_report_by_keyword(product_keyword : str, db) -> Tuple[io.BytesIO, str]:
    products = ProductRepository.get_by_keyword(db, product_keyword)
    if products is None:
        raise KeyError(f"Product containing {product_keyword} does not exist!")

    workbook = Workbook()
    workbook_state = workbook.active
    workbook_state.title = f"Отчёт по товару {product_keyword}"

    workbook_state.append([
        "Продавец", "Цена по Ozon карте", "Цена", "Рейтинг", "Количество отзывов",
        "Количество вопросов", "На распродаже", "Дата", "Ссылка(товар)", "Ссылка(продавец)"
    ])

    for product in products:
        seller = SellerRepository.get_by_id(db, product.seller_id)
        state = ProductRepository.get_last_state(db, product.id)
        seller_name = seller.name
        price_ozon = state.price_ozon_card
        price = state.price
        rating = state.rating
        reviews = state.review_count
        questions = state.question_count
        on_sale = state.on_sale
        created_at = state.created_at
        product_url = product.url
        seller_url = seller.url
        workbook_state.append([
            seller_name, price_ozon, price, rating, reviews, questions,
            on_sale, created_at, product_url, seller_url
        ])

    report_file = io.BytesIO()
    workbook.save(report_file)
    report_file.seek(0)
    file_name = f"{product_keyword}.xlsx"
    return report_file, file_name
